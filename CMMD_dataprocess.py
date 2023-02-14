import openpyxl
import cv2 
import numpy as np
import os,shutil
from openpyxl import Workbook

np.set_printoptions(threshold=50000)

def LorR(img):  #determine whether the picture is left breast or by breast
    array = np.array(img)
    pixel = array[:,0]

    if(np.any(pixel)):
        return "L"
    else:
        return "R"

data = openpyxl.load_workbook("****.xlsx")  #calcification image information of all patients

table = data.get_sheet_by_name("Sheet1")

IDs = []
ID1s = []
LeftRights = []
Ages = []

for col in table.iter_cols(min_col=1,max_col=1):
    for cell in col:
        value = cell.value
        IDs.append(value)

for col in table.iter_cols(min_col=2,max_col=2):
    for cell in col:
        value = cell.value
        ID1s.append(value)

for col in table.iter_cols(min_col=3,max_col=3):
    for cell in col:
        value = cell.value
        LeftRights.append(value)

for col in table.iter_cols(min_col=4,max_col=4):
    for cell in col:
        value = cell.value
        Ages.append(value)

print(len(IDs),len(ID1s),len(LeftRights))

data2 = openpyxl.load_workbook(filename = "****.xlsx")  #record the patient's normal picture
sheet = data2.active
line = 2

for i in range(1,1872):
    ID = IDs[i]
    ID1 = ID1s[i]
    LeftRight = LeftRights[i]
    Age = Ages[i]
    if(ID == IDs[i+1] or ID == IDs[i-1]):
        continue
    
    print(ID1)

    path = "./alldata/"+ID+"/"
    filelist = os.listdir(path)

    if(not os.path.exists('./CMMD3/'+ID1)):
        os.makedirs('./CMMD3/'+ID1)

    num=1
    for name in filelist:
        if(len(name.split("."))==1 or name.split(".")[1]!="tif"):
            continue
        img = cv2.imdecode(np.fromfile(path+name,dtype=np.uint8),-1)
        pos = LorR(img)
        if(pos!=LeftRight):
            shutil.copy(path+name, './CMMD3/'+ID1+'/'+pos+str(num)+'.tif')
            num+=1
    sheet['A'+str(line)] = ID1
    if(LeftRight=='L'):
        sheet['B'+str(line)] = 'R'
    else:
        sheet['B'+str(line)] = 'L'
    sheet['C'+str(line)] = Age
    sheet['D'+str(line)] = num-1
    sheet['E'+str(line)] = 'normal'
    line +=1

data2.save(filename='****.xlsx')  #save the patient's normal picture

    




